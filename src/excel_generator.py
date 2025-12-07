"""
Excel Generator Module for SCAG I-5 Analysis

This module generates professional Excel reports with multiple sheets:
- Summary sheet with overall statistics
- Detailed analysis sheets for AADT, Peak Hours, Capacity, and Trucks
- Professional formatting: fonts, colors, borders, number formats
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
import logging

from . import config
from .utils import log_analysis_step

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generator for professional Excel reports with formatting.

    This class creates multi-sheet Excel reports with:
    1. Summary sheet - Overall statistics for all segments
    2. Detailed analysis sheets - AADT, Peak Hours, Capacity, Truck Analysis
    3. Professional formatting - Fonts, colors, borders, number formats

    Attributes:
        workbook (Workbook): openpyxl Workbook object
        output_path (str): Path to save the Excel file

    Styling Standards:
        - Header: Bold, White text on Dark Blue background (#366092)
        - Data: Calibri 11pt, with appropriate number formats
        - Borders: Thin borders around all cells
        - Alignment: Center for headers, left for text, right for numbers
    """

    def __init__(self, output_path: str) -> None:
        """
        Initialize the Excel Generator.

        Args:
            output_path: Path where the Excel file will be saved

        Example:
            >>> generator = ExcelGenerator('output/report.xlsx')
        """

        self.wb = Workbook()
        self.output_path = output_path
        self.wb.remove(self.wb.active)  # Remove default sheet if needed.

    def _create_header_style(self) -> Dict:
        """
        Create standard header style for all sheets.


        Returns:
            dict: Dictionary containing Font, PatternFill, Border, Alignment objects
                {
                    'font': Font(...),
                    'fill': PatternFill(...),
                    'border': Border(...),
                    'alignment': Alignment(...)
                }

        Example:
            >>> generator = ExcelGenerator('output.xlsx')
            >>> header_style = generator._create_header_style()
            >>> ws['A1'].font = header_style['font']
            >>> ws['A1'].fill = header_style['fill']

        Steps to implement:
            1. Create Font: Calibri 12pt, Bold, White color
            2. Create PatternFill: Dark blue (#366092), solid fill
            3. Create Border: Thin borders on all sides
            4. Create Alignment: Center horizontal and vertical
            5. Return dictionary with all style objects
        """

        style_dict = {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            ),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }
        return style_dict

    def _apply_cell_style(
        self,
        cell,
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None,
        border: Optional[Border] = None,
        alignment: Optional[Alignment] = None,
        number_format: Optional[str] = None,
    ) -> None:
        """
        Apply styling to a single cell.

        Args:
            cell: openpyxl Cell object
            font: Font style (optional)
            fill: Fill style (optional)
            border: Border style (optional)
            alignment: Alignment style (optional)
            number_format: Number format string (optional)

        Example:
            >>> generator._apply_cell_style(
            ...     ws['A1'],
            ...     font=Font(bold=True),
            ...     fill=PatternFill(start_color='366092', fill_type='solid'),
            ...     alignment=Alignment(horizontal='center')
            ... )

        Steps to implement:
            1. Check if font is provided, apply to cell.font
            2. Check if fill is provided, apply to cell.fill
            3. Check if border is provided, apply to cell.border
            4. Check if alignment is provided, apply to cell.alignment
            5. Check if number_format is provided, apply to cell.number_format
        """

        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format

    def _set_column_widths(self, worksheet, widths: Dict[str, float]) -> None:
        """
        Set column widths for a worksheet.

        Args:
            worksheet: openpyxl Worksheet object
            widths: Dictionary mapping column letters to widths
                    e.g., {'A': 20, 'B': 15, 'C': 12}

        Example:
            >>> generator._set_column_widths(ws, {'A': 20, 'B': 15, 'C': 12})

        Steps to implement:
            1. Loop through widths dictionary items
            2. For each column letter and width
            3. Set worksheet.column_dimensions[column_letter].width = width
        """
        # TODO: Implement this method
        # Hint: Use worksheet.column_dimensions
        for key, value in widths.items():
            worksheet.column_dimensions[key].width = value

    def create_summary_sheet(
        self, summary_data: pd.DataFrame, sheet_name: str = "Summary_all"
    ) -> None:
        """
        Create summary sheet with overall statistics.

        Args:
            summary_data: DataFrame containing summary statistics
            sheet_name: Name of the sheet (default: "Summary_all")

        Expected DataFrame columns:
            - Year, Section, Direction, Type
            - AADT, Truck_PCT, Peak_AM, Peak_PM
            - VC_Ratio_AM, VC_Ratio_PM, LOS_AM, LOS_PM

        Example:
            >>> summary_df = pd.DataFrame({...})
            >>> generator.create_summary_sheet(summary_df)

        Steps to implement:
            1. Log the start of sheet creation
            2. Create new worksheet with sheet_name
            3. Get header style from _create_header_style()
            4. Write headers (column names) to first row
            5. Apply header style to first row
            6. Write data rows (from DataFrame)
            7. Apply data formatting:
               - AADT columns: '#,##0' (integer with thousands separator)
               - Percentage columns: '0.0%'
               - VC Ratio columns: '0.00'
            8. Apply borders to all data cells
            9. Set column widths appropriately
            10. Freeze first row (freeze_panes = 'A2')
            11. Log completion
        """
        # TODO: Implement this method
        # Hint 1: Use ws.append() to write rows
        # Hint 2: Iterate through cells to apply formatting
        # Hint 3: Use dataframe.values.tolist() to get data rows
        log_analysis_step("Excel Generator", "Start creating summary sheet")

        ws = self.wb.create_sheet(sheet_name)
        header_style = self._create_header_style()
        ws.append(summary_data.columns)
        for i in range(len(summary_data.columns)):
            self._apply_cell_style(
                ws.cell(row=1, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = summary_data.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells
        num_data_rows = len(summary_data)
        for row_idx in range(2, 2 + num_data_rows):
            for col_idx in range(len(summary_data.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = summary_data.columns[col_idx]

                if "AADT" in column_name or "Peak" in column_name:
                    cell.number_format = "#,##0"
                elif "PCT" in column_name:
                    cell.number_format = "0.0%"
                elif "VC_Ratio" in column_name:
                    cell.number_format = "0.00"

        # set the column widths
        column_widths = {
            "A": 8,  # Year
            "B": 12,  # Section
            "C": 12,  # Direction
            "D": 18,  # Type
            "E": 15,  # AADT
            "F": 12,  # Truck_PCT
            "G": 12,  # Peak_AM
            "H": 12,  # Peak_PM
            "I": 13,  # VC_Ratio_AM
            "J": 13,  # VC_Ratio_PM
            "K": 10,  # LOS_AM
            "L": 10,  # LOS_PM
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A2"

        log_analysis_step("Excel Generator", "Completed creating summary sheet.")

    def create_aadt_sheet(self, aadt_data: pd.DataFrame) -> None:
        """
        Create AADT analysis sheet.

        Args:
            aadt_data: DataFrame containing AADT analysis results

        Expected columns:
            - Direction, Type, Num_Segments
            - Avg_AADT, Min_AADT, Max_AADT
            - Avg_Auto_AADT, Avg_Truck_AADT, Avg_Truck_PCT

        Example:
            >>> aadt_df = aadt_calculator.calculate_all_groups_aadt()
            >>> generator.create_aadt_sheet(aadt_df)

        Steps to implement:
            1. Log the start
            2. Create worksheet named "AADT_Analysis"
            3. Write title row: "AADT Analysis by Direction and Facility Type"
            4. Apply title formatting (larger font, bold, merged cells)
            5. Write headers
            6. Write data with appropriate number formats:
               - AADT values: '#,##0'
               - Truck PCT: '0.0%'
            7. Apply styling and borders
            8. Set column widths
            9. Freeze panes
            10. Log completion
        """
        # TODO: Implement this method
        # Hint: Similar to create_summary_sheet() but with different columns

        log_analysis_step("Excel Generator", "Start creating AADT sheet")

        ws = self.wb.create_sheet("AADT_Analysis")

        # create a title
        ws["A1"] = "AADT Analysis by Direction and Facility Type"
        ws.merge_cells("A1:I1")

        # title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        # header formatting
        header_style = self._create_header_style()
        ws.append(aadt_data.columns)
        for i in range(len(aadt_data.columns)):
            self._apply_cell_style(
                ws.cell(row=2, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = aadt_data.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells
        num_data_rows = len(aadt_data)
        for row_idx in range(3, 3 + num_data_rows):
            for col_idx in range(len(aadt_data.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = aadt_data.columns[col_idx]

                if "AADT" in column_name or "Peak" in column_name:
                    cell.number_format = "#,##0"
                elif "PCT" in column_name:
                    cell.number_format = "0.0%"

                self._apply_cell_style(cell, border=header_style["border"])

        # set the column widths
        column_widths = {
            "A": 12,  # Direction
            "B": 18,  # Type
            "C": 12,  # Num_Segments
            "D": 15,  # Avg_AADT
            "E": 15,  # Min_AADT
            "F": 15,  # Max_AADT
            "G": 15,  # Avg_Auto_AADT
            "H": 15,  # Avg_Truck_AADT
            "I": 15,  # Avg_Truck_PCT
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A3"

        log_analysis_step("Excel Generator", "Completed creating aadt sheet.")

    def create_peak_hour_sheet(self, peak_data: pd.DataFrame) -> None:
        """
        Create peak hour analysis sheet.

        Args:
            peak_data: DataFrame containing peak hour analysis results

        Expected columns:
            - Direction, Type, Period
            - Avg_Peak_Total, Avg_Peak_Auto, Avg_Peak_Truck
            - Min_Peak_Total, Max_Peak_Total

        Example:
            >>> am_peaks = peak_analyzer.calculate_all_groups_peak('AM')
            >>> generator.create_peak_hour_sheet(am_peaks)
        """
        # TODO: Implement this method
        # Hint: Follow same pattern as create_aadt_sheet()

        log_analysis_step("Excel Generator", "Start creating peak hour sheet")

        ws = self.wb.create_sheet("Peak_Hour_Analysis")

        # create a title
        ws["A1"] = "Peak Hour Analysis by Period"
        ws.merge_cells("A1:H1")

        # title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        # header formatting
        header_style = self._create_header_style()
        ws.append(peak_data.columns)
        for i in range(len(peak_data.columns)):
            self._apply_cell_style(
                ws.cell(row=2, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = peak_data.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells
        num_data_rows = len(peak_data)
        for row_idx in range(3, 3 + num_data_rows):
            for col_idx in range(len(peak_data.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = peak_data.columns[col_idx]

                if "Peak" in column_name:
                    cell.number_format = "#,##0"
                elif "PCT" in column_name:
                    cell.number_format = "0.0%"

                self._apply_cell_style(cell, border=header_style["border"])

        # set the column widths
        column_widths = {
            "A": 12,  # Direction
            "B": 18,  # Type
            "C": 12,  # Period
            "D": 15,  # Avg_Peak_Total
            "E": 15,  # Avg_Peak_Auto
            "F": 15,  # Avg_Peak_Truck
            "G": 15,  # Min_Peak_Total
            "H": 15,  # Max_Peak_Total
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A3"

        log_analysis_step(
            "Excel Generator", "Completed creating peak hour analysis sheet."
        )

    def create_capacity_sheet(self, capacity_data: pd.DataFrame) -> None:
        """
        Create capacity and LOS analysis sheet.

        Args:
            capacity_data: DataFrame containing capacity analysis results

        Expected columns:
            - Direction, Type, Period
            - Avg_PCE_Flow, Avg_Capacity, Avg_VC_Ratio
            - Dominant_LOS, LOS_Counts

        Example:
            >>> am_capacity = capacity_analyzer.calculate_all_groups_capacity('AM')
            >>> generator.create_capacity_sheet(am_capacity)

        Special formatting:
            - VC Ratio: '0.00' format
            - LOS grades: Center alignment
            - Color coding for LOS levels:
              - A, B: Green
              - C, D: Yellow
              - E, F: Red
        """
        # TODO: Implement this method
        # Hint: Add conditional formatting for LOS grades
        log_analysis_step("Excel Generator", "Start creating capacity sheet")

        ws = self.wb.create_sheet("Capacity_Analysis")

        # create a title
        ws["A1"] = "Capacity Analysis by Period"
        ws.merge_cells("A1:H1")

        # title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        # header formatting
        header_style = self._create_header_style()
        ws.append(capacity_data.columns)
        for i in range(len(capacity_data.columns)):
            self._apply_cell_style(
                ws.cell(row=2, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = capacity_data.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells
        num_data_rows = len(capacity_data)
        for row_idx in range(3, 3 + num_data_rows):
            for col_idx in range(len(capacity_data.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = capacity_data.columns[col_idx]

                if "VC" in column_name:
                    cell.number_format = "0.00"
                elif "LOS" in column_name:
                    cell.alignment = header_style["alignment"]
                    if cell.value in ["A", "B"]:
                        cell.fill = PatternFill(
                            start_color="90EE90",
                            end_color="90EE90",
                            fill_type="solid",
                        )
                    elif cell.value in ["C", "D"]:
                        cell.fill = PatternFill(
                            start_color="FFFF00",
                            end_color="FFFF00",
                            fill_type="solid",
                        )
                    elif cell.value in ["E", "F"]:
                        cell.fill = PatternFill(
                            start_color="FF6347",
                            end_color="FF6347",
                            fill_type="solid",
                        )

                self._apply_cell_style(cell, border=header_style["border"])

        # set the column widths
        column_widths = {
            "A": 12,  # Direction
            "B": 18,  # Type
            "C": 12,  # Period
            "D": 15,  # Avg_PCE_Flow
            "E": 18,  # Avg_Capacity
            "F": 15,  # Avg_VC_Ratio
            "G": 12,  # Dominant_LOS
            "H": 12,  # LOS_Counts
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A3"

        log_analysis_step(
            "Excel Generator", "Completed creating capacity analysis sheet."
        )

    def create_truck_sheet(self, truck_data: pd.DataFrame) -> None:
        """
        Create truck analysis sheet.

        Args:
            truck_data: DataFrame containing truck analysis results

        Expected columns:
            - Direction, Type
            - Avg_Truck_AADT, Avg_Truck_PCT, Avg_Truck_Intensity
            - Avg_AM_Truck_Ratio, Avg_PM_Truck_Ratio
            - Min_Truck_PCT, Max_Truck_PCT

        Example:
            >>> truck_df = truck_analyzer.calculate_all_groups_truck()
            >>> generator.create_truck_sheet(truck_df)
        """
        # TODO: Implement this method
        # Hint: Follow same pattern as other sheet creation methods
        log_analysis_step("Excel Generator", "Start creating truck analysis sheet")

        ws = self.wb.create_sheet("Truck_Analysis")

        # create a title
        ws["A1"] = "Truck Analysis"
        ws.merge_cells("A1:I1")

        # title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        # header formatting
        header_style = self._create_header_style()
        ws.append(truck_data.columns)
        for i in range(len(truck_data.columns)):
            self._apply_cell_style(
                ws.cell(row=2, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = truck_data.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells

        num_data_rows = len(truck_data)
        for row_idx in range(3, 3 + num_data_rows):
            for col_idx in range(len(truck_data.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = truck_data.columns[col_idx]

                if "AADT" in column_name:
                    cell.number_format = "#,##0"
                elif "PCT" in column_name:
                    cell.number_format = "0.0%"
                elif "Ratio" in column_name or "Intensity" in column_name:
                    cell.number_format = "0.00"

                self._apply_cell_style(cell, border=header_style["border"])

        # set the column widths
        column_widths = {
            "A": 12,  # Direction
            "B": 18,  # Type
            "C": 15,  # Avg_Truck_AADT
            "D": 15,  # Avg_Truck_PCT
            "E": 15,  # Avg_Truck_Intensity
            "F": 15,  # Avg_AM_Truck_Ratio
            "G": 15,  # Avg_PM_Truck_Ratio
            "H": 15,  # Min_Truck_PCT
            "I": 15,  # Max_Truck_PCT
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A3"

        log_analysis_step("Excel Generator", "Completed creating truck analysis sheet.")

    def create_comparison_sheet(
        self, am_data: pd.DataFrame, pm_data: pd.DataFrame
    ) -> None:
        """
        Create AM vs PM comparison sheet.

        Args:
            am_data: DataFrame with AM period data
            pm_data: DataFrame with PM period data

        This sheet shows side-by-side comparison of AM and PM metrics:
            - Peak flows
            - VC ratios
            - LOS grades
            - Truck percentages

        Example:
            >>> am_df = capacity_analyzer.calculate_all_groups_capacity('AM')
            >>> pm_df = capacity_analyzer.calculate_all_groups_capacity('PM')
            >>> generator.create_comparison_sheet(am_df, pm_df)

        Steps to implement:
            1. Create worksheet "AM_vs_PM_Comparison"
            2. Merge AM and PM data on Direction and Type
            3. Create headers for both periods side by side
            4. Write merged data
            5. Add calculated columns:
               - Difference (AM - PM)
               - Higher period indicator
            6. Apply formatting with period headers in different colors
            7. Highlight differences (conditional formatting)
        """
        # TODO: Implement this method
        # Hint: Use pd.merge() to combine AM and PM data
        # Hint: Use different fill colors for AM (blue) and PM (orange) headers
        log_analysis_step("Excel Generator", "Start creating period comparison sheet")

        ws = self.wb.create_sheet("AM_vs_PM_Comparison")

        merged_df = pd.merge(
            am_data, pm_data, on=["Direction", "Type"], suffixes=("_AM", "_PM")
        )

        if "Avg_Peak_Total_AM" in merged_df.columns:
            merged_df["Peak_Diff"] = (
                merged_df["Avg_Peak_Total_AM"] - merged_df["Avg_Peak_Total_PM"]
            ).abs()

        if "Avg_VC_Ratio_AM" in merged_df.columns:
            merged_df["VC_Ratio_Diff"] = (
                merged_df["Avg_VC_Ratio_AM"] - merged_df["Avg_VC_Ratio_PM"]
            ).abs()

        if "Avg_Truck_PCT_AM" in merged_df.columns:
            merged_df["Truck_PCT_Diff"] = (
                merged_df["Avg_Truck_PCT_AM"] - merged_df["Avg_Truck_PCT_PM"]
            ).abs()

        # title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # create a title
        ws["A1"] = "AM and PM comparison metrics"
        num_cols = len(merged_df.columns)
        last_col = get_column_letter(num_cols)
        ws.merge_cells(f"A1:{last_col}1")

        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        # header formatting
        header_style = self._create_header_style()
        ws.append(merged_df.columns)
        for i in range(len(merged_df.columns)):
            self._apply_cell_style(
                ws.cell(row=2, column=i + 1),
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # Write data rows
        df_list = merged_df.values.tolist()
        for row in df_list:
            ws.append(row)

        # Apply data formatting and apply borders to all data cells

        # los color settings
        los_colors = {
            "A": "90EE90",  # Light green
            "B": "ADFF2F",  # Green yellow
            "C": "FFFF00",  # Yellow
            "D": "FFA500",  # Orange
            "E": "FF6347",  # Tomato
            "F": "FF0000",  # Red
        }

        num_data_rows = len(merged_df)
        for row_idx in range(3, 3 + num_data_rows):
            for col_idx in range(len(merged_df.columns)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                column_name = merged_df.columns[col_idx]

                if "Flow" in column_name:
                    cell.number_format = "#,##0"
                elif "PCT" in column_name:
                    cell.number_format = "0.0%"
                elif "Ratio" in column_name or "Intensity" in column_name:
                    cell.number_format = "0.00"
                elif "Dominant_LOS" in column_name:
                    cell.alignment = header_style["alignment"]
                    if cell.value in los_colors:
                        cell.fill = PatternFill(
                            start_color=los_colors[cell.value],
                            end_color=los_colors[cell.value],
                            fill_type="solid",
                        )

                self._apply_cell_style(cell, border=header_style["border"])

        # set the column widths

        column_widths = {
            "A": 12,  # Direction
            "B": 18,  # Type
            "C": 15,  # Avg_PCE_Flow_AM
            "D": 15,  # Avg_PCE_Flow_PM
            "E": 15,  # Peak_PCE_Flow_Diff
            "F": 15,  # Avg_VC_ratios_AM
            "G": 15,  # Avg_VC_ratios_PM
            "H": 12,  # VC_Ratio_Diff
            "I": 12,  # Dominant_LOS_AM
            "J": 12,  # Dominant_LOS_PM
            "K": 12,  # LOS_Counts_AM
            "L": 12,  # LOS_Counts_PM
        }

        self._set_column_widths(ws, column_widths)

        # freeze first row
        ws.freeze_panes = "A3"

        log_analysis_step(
            "Excel Generator", "Completed creating AM and PM comparison sheet."
        )

    def add_metadata_sheet(self, metadata: Dict) -> None:
        """
        Add metadata sheet with analysis information.

        Args:
            metadata: Dictionary containing analysis metadata
                {
                    'analysis_date': '2025-12-02',
                    'model': 'SCAG 2024 ABM',
                    'analyst': 'Your Name',
                    'base_year': 2019,
                    'forecast_year': 2045,
                    'sections_analyzed': 3,
                    'total_segments': 61
                }

        Example:
            >>> metadata = {
            ...     'analysis_date': '2025-12-02',
            ...     'model': 'SCAG 2024 ABM'
            ... }
            >>> generator.add_metadata_sheet(metadata)

        Steps to implement:
            1. Create worksheet "Metadata"
            2. Write metadata as key-value pairs
            3. Apply formatting to make it readable
            4. Add project description at top
            5. Include data source information
        """
        log_analysis_step("Excel Generator", "Start adding metadata sheet")

        ws = self.wb.create_sheet("Metadata")
        header_style = self._create_header_style()

        ws["A1"] = "Analysis Metadata"
        ws.merge_cells("A1:B1")

        # Title formatting
        title_font = Font(name="Calibri", size=16, bold=True)
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = title_border

        ws["A3"] = "Key"
        ws["B3"] = "Value"

        for col in ["A", "B"]:
            cell = ws[f"{col}3"]
            self._apply_cell_style(
                cell,
                font=header_style["font"],
                fill=header_style["fill"],
                border=header_style["border"],
                alignment=header_style["alignment"],
            )

        # fill in metadata from row 4
        row_num = 4
        for key, value in metadata.items():
            ws.cell(row=row_num, column=1, value=key)
            ws.cell(row=row_num, column=2, value=value)

            for col_idx in [1, 2]:
                self._apply_cell_style(
                    ws.cell(row=row_num, column=col_idx), border=header_style["border"]
                )
            row_num += 1

        column_widths = {
            "A": 25,  # key
            "B": 40,  # value
        }

        self._set_column_widths(ws, column_widths)

        log_analysis_step("Excel Generator", "Completed adding metadata sheet")

    def save(self) -> None:
        """
        Save the workbook to file.

        Example:
            >>> generator.save()

        Steps to implement:
            1. Log the save operation
            2. Save workbook to self.output_path
            3. Log completion with file path
            4. Handle any save errors
        """
        try:
            log_analysis_step("Excel Generator", f"Saving Excel to {self.output_path}")
            self.wb.save(self.output_path)
            log_analysis_step("Excel Generator", f"Saved Excel to {self.output_path}")
        except PermissionError as e:
            logger.error(
                f"Permission denied when saving to {self.output_path}: {str(e)}"
            )
            raise
        except OSError as e:
            logger.error(f"Failed to save workbook to {self.output_path}: {str(e)}")
            raise
